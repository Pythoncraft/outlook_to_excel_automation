import win32com.client as client
import re
import os
import openpyxl as xl

outlook = client.Dispatch("Outlook.Application")
namespace = outlook.GetNamespace("MAPI")

inbox = namespace.GetDefaultFolder(6) # "6" refers to the inbox folder. returns an object

leads_folder = inbox.Folders['Leads'] # refering / accessing the 'Leads' subfolder
messages = leads_folder.items

wb = xl.load_workbook('MPRnew.xlsx')
ws = wb['Lead Data']
Sender_Contact = []
for msg in messages:
	if msg.Class==43: # Class 43 is a MAILITEM
		if msg.SenderEmailType=='EX': # checking if the message type is EXCHANGE
			ws['c6'] = msg.Sender.GetExchangeUser().PrimarySmtpAddress
		else:
			ws['c6'] = msg.SenderEmailAddress
		ws['b6'] = str(msg.SenderName)
		ws['d6'] = 'Google/Website'
		ws.insert_rows(6)
	else:
		print(f'Different class mail from'+ msg.SenderName)

wb.save('MPRnew.xlsx')

# ws['b6'] = str(message.SenderName)
# ws['c6'] = Sender_Contact
# ws['d6'] = 'Google/Website'
# ws['e6'] = str(message.CreationTime)
# ws['f6'] = str(message.body)
# ws.insert_rows(6)


#######Functions/Methods#############
# print(message.CreationTime)
# message.display() - opens outllook message to view
# item.Subject
# item.Body
# item.SenderName
# item.Class -- check item class Mailitem / ReportItem, MeetingItem etc..
# item.SenderEmailType -- check items type SMTP Exchange or other
# item.SenderEmailAddress -- only for 'SMTP' type emails
# item.Sender.GetExchangeUser().PrimarySmtpAddress -- for EXCHANGE type emails 

########## INSERT DATA TO EXCEL FILE ##############

# CREATE DIRECTORY
# os.mkdir(os.path.join(os.getcwd(), folder_name))

# SAVE / WRITE FILES
# name_only = links.split('/')[-1] # splitting the link text to use as file name
# 		with open(name_only + '.jpg', 'wb') as f: # open the file in wirte mode 'w', # 'wb' is the 'write binary' mode
# 			im = requests.get(links) # sending individual requests to the links, to get the informatio from them
# 			f.write(im.content) # write and save content (information) that the link contains  to file

